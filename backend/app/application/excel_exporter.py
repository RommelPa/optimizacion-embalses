from __future__ import annotations

import os
import tempfile
from io import BytesIO
from typing import cast

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.application.corrida_resultados_builder import build_resultados_dataset
from app.models.corrida import Corrida
from app.application.corrida_chart_builder import (
    render_caudal_chart,
    render_despacho_chart,
    render_volumenes_chart,
)

def insertar_imagen_excel(
    ws: Worksheet,
    image_path: str,
    anchor: str,
    scale: float = 0.55,
) -> None:
    img = Image(image_path)
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    img.anchor = anchor
    ws.add_image(img)


def build_excel_corrida_legacy(corrida: Corrida) -> tuple[bytes, str]:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Resultados"

    dataset = build_resultados_dataset(corrida)

    meta = dataset["meta"]
    tabla_rows = dataset["tabla"]["rows"]
    caudal = dataset["caudal"]
    volumenes = dataset["volumenes"]
    despacho = dataset["despacho"]
    validacion = dataset["validacion"]

    horas = meta["horas"]
    periodos = np.array(meta["periodos"], dtype=np.int32)
    horas_etiquetas = meta["horas_etiquetas"]
    etiquetas_todas = meta["etiquetas_todas"]
    tick_step = meta["tick_step"]
    titulo_modo = meta["titulo_modo"]
    q_salida_campanario = float(meta["q_salida_campanario"])
    correlacion = float(meta["correlacion"])
    ingresos_totales = float(meta["ingresos_totales"])

    q_opt_arr = np.array(caudal["q_opt"], dtype=np.float64)
    q_cincel_arr = np.array(caudal["q_entrada_cincel"], dtype=np.float64)
    q_rango_min = float(caudal["q_rango_min"])
    q_rango_max = float(caudal["q_rango_max"])

    v_cincel_arr = np.array(volumenes["v_cincel"], dtype=np.float64)
    v_camp_arr = np.array(volumenes["v_campanario"], dtype=np.float64)
    v_cincel_min = float(volumenes["v_cincel_min"])
    v_cincel_max = float(volumenes["v_cincel_max"])
    v_camp_min = float(volumenes["v_campanario_min"])
    v_camp_max = float(volumenes["v_campanario_max"])

    cmg_arr = np.array(despacho["cmg"], dtype=np.float64)
    pot_ch4_arr = np.array(despacho["potencia_ch4"], dtype=np.float64)
    pot_ch6_arr = np.array(despacho["potencia_ch6"], dtype=np.float64)

    viol_cincel_sobre = sum(1 for x in v_cincel_arr if x > v_cincel_max)
    viol_cincel_bajo = sum(1 for x in v_cincel_arr if x < v_cincel_min)
    viol_camp_sobre = sum(1 for x in v_camp_arr if x > v_camp_max)
    viol_camp_bajo = sum(1 for x in v_camp_arr if x < v_camp_min)
    viol_camp_neg = sum(1 for x in v_camp_arr if x < 0)

    es_valida = (
        viol_cincel_sobre == 0
        and viol_cincel_bajo == 0
        and viol_camp_sobre == 0
        and viol_camp_bajo == 0
        and viol_camp_neg == 0
    )

    headers = [
        "HORA",
        "P_Char 5 (MW)",
        "Volumen D. Cincel (m3)",
        "Caudal Salida D. Cincel (m3/s)",
        "Volumen D. Campanario (m3)",
        "Caudal salida D. Campanario (m3/s)",
        "CH 4 (MW)",
        "CH 6 (MW)",
        "CH 123 (MW)",
        "CMG (S/./MWh)",
        "Ingresos",
    ]
    ws.append(headers)

    for row in tabla_rows:
        ws.append(
            [
                row["hora"],
                row["p_char_5"],
                int(row["v_cincel"]) if row["v_cincel"] is not None else None,
                round(row["q_opt"], 1) if row["q_opt"] is not None else None,
                int(row["v_campanario"]) if row["v_campanario"] is not None else None,
                round(row["q_salida_campanario"], 1)
                if row["q_salida_campanario"] is not None
                else None,
                round(row["potencia_ch4"], 1)
                if row["potencia_ch4"] is not None
                else None,
                round(row["potencia_ch6"], 1)
                if row["potencia_ch6"] is not None
                else None,
                row["ch123"],
                round(row["cmg"], 1) if row["cmg"] is not None else None,
                round(row["ingreso"], 2) if row["ingreso"] is not None else None,
            ]
        )

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F497D")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, 12):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    anchos = [8, 14, 22, 28, 26, 32, 10, 10, 12, 16, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for row_idx in range(2, horas + 2):
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_align
            cell.border = thin_border

    fila_resumen = horas + 3
    resumen_datos = [
        ("Caso de estudio", corrida.caso_estudio),
        ("Modo", titulo_modo),
        ("Ingresos totales período reopt ($)", round(ingresos_totales, 2)),
        ("Tiempo optimización (min)", round(corrida.execution_time_sec / 60.0, 3)),
        ("Tiempo total (min)", round(corrida.execution_time_sec / 60.0, 3)),
        ("Correlación CMG-Potencia total", round(correlacion, 4)),
        (
            "Q promedio optimizado (m³/s)",
            round(float(np.mean(q_opt_arr)), 3) if len(q_opt_arr) > 0 else 0,
        ),
        (
            "Vol. D. Cincel inicial (mil m³)",
            round(validacion["v_cincel_inicial"] / 1000, 1),
        ),
        (
            "Vol. D. Cincel final (mil m³)",
            round(validacion["v_cincel_final"] / 1000, 1),
        ),
        (
            "Vol. D. Campanario inicial (mil m³)",
            round(validacion["v_campanario_inicial"] / 1000, 1),
        ),
        (
            "Vol. D. Campanario final (mil m³)",
            round(validacion["v_campanario_final"] / 1000, 1),
        ),
        ("Solución válida", "SÍ" if es_valida else "NO - REVISAR"),
    ]

    for idx, (etiqueta, valor) in enumerate(resumen_datos):
        ws.cell(row=fila_resumen + idx, column=1, value=etiqueta)
        ws.cell(row=fila_resumen + idx, column=2, value=valor)

    with tempfile.TemporaryDirectory() as tmpdir:
        ruta_caudal = os.path.join(tmpdir, "grafico_caudal.png")
        ruta_cmg = os.path.join(tmpdir, "grafico_cmg_potencia.png")
        ruta_vol = os.path.join(tmpdir, "grafico_volumenes.png")

        fig1, ax1 = plt.subplots(figsize=(10, 5))
        render_caudal_chart(ax1, dataset)
        fig1.tight_layout()
        fig1.savefig(ruta_caudal, dpi=150, bbox_inches="tight")
        plt.close(fig1)

        fig2, ax2 = plt.subplots(figsize=(10, 5))
        render_despacho_chart(ax2, dataset)
        fig2.tight_layout()
        fig2.savefig(ruta_cmg, dpi=150, bbox_inches="tight")
        plt.close(fig2)

        fig3, ax3 = plt.subplots(figsize=(10, 5))
        render_volumenes_chart(ax3, dataset)
        fig3.tight_layout()
        fig3.savefig(ruta_vol, dpi=150, bbox_inches="tight")
        plt.close(fig3)

        insertar_imagen_excel(ws, ruta_caudal, "M2")
        insertar_imagen_excel(ws, ruta_cmg, "M25")
        insertar_imagen_excel(ws, ruta_vol, "M48")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Resultados_PSO_CH46_Q{int(q_salida_campanario)}.xlsx"
        return output.getvalue(), filename