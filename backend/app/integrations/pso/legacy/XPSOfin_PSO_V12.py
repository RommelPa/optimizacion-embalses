
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import pyswarms as ps
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from scipy.ndimage import uniform_filter1d
import time


# =============================================
# CONFIGURACIÓN INICIAL
# =============================================

start_time = time.time()

# Parámetros del sistema
PERIODOS_TOTALES = 48
rendimiento_CH4 = 1.01   # Charcani 4 (más cerca al Dique Cincel)
rendimiento_CH6 = 0.59   # Charcani 6 (más cerca al Dique Campanario)

# Límites de embalses (Dique Cincel y Dique Campanario)
V_Cincel_max,     V_Cincel_min     = 190000.0, 20000.0
V_Campanario_max, V_Campanario_min =  90000.0, 20000.0

# Rango operativo del caudal
Q_rango = [6.0, 15.0]


# =============================================
# MODO DE OPERACIÓN: PROGRAMA INICIAL o REPROGRAMA
# =============================================

def hora_a_periodo(hora_str):
    """Convierte 'HH:MM' a índice de periodo (0-based). Ej: '00:30' → 0, '10:00' → 19"""
    partes = hora_str.strip().split(':')
    h = int(partes[0])
    m = int(partes[1])
    minutos = h * 60 + m
    periodo = minutos // 30 - 1   # periodo 0-based: 00:30 = periodo 0
    return periodo

def periodo_a_hora(periodo):
    """Convierte índice de periodo (0-based) a string 'HH:MM'. Ej: 0 → '00:30'"""
    minutos = (periodo + 1) * 30
    h = (minutos // 60) % 24
    m = minutos % 60
    if h == 0 and m == 0:
        return "24:00"
    return f"{h:02d}:{m:02d}"

print("=" * 70)
print("  SISTEMA DE OPTIMIZACIÓN PSO — CHARCANI 4 & 6")
print("=" * 70)
print()
print("Seleccione el modo de operación:")
print("  [1] Programa Inicial  (optimización completa 00:30 → 24:00)")
print("  [2] Reprogramación    (re-optimización desde una hora de corte)")
print()

while True:
    modo_input = input("Ingrese opción (1 o 2): ").strip()
    if modo_input in ('1', '2'):
        break
    print("  → Opción no válida. Ingrese 1 o 2.")

ES_REPROGRAMA = (modo_input == '2')

# Variables de reprogramación (se llenan si ES_REPROGRAMA)
periodo_corte  = 0       # índice 0-based del primer periodo RE-optimizado
hora_corte_str = "00:00"
V_Cincel_corte     = None
V_Campanario_corte = None
Q_salida_Campanario_nuevo = None

# Datos del programa anterior (para graficar en gris)
Q_opt_anterior      = None
V_Cincel_anterior   = None
V_camp_anterior     = None
CMG_anterior        = None
Pot_entrada_anterior = None

if ES_REPROGRAMA:
    print()
    print("─" * 50)
    print("  CONFIGURACIÓN DE REPROGRAMACIÓN")
    print("─" * 50)

    # ── Hora de corte ──────────────────────────────────────
    print()
    print("Ingrese la hora de corte (primer período RE-optimizado).")
    print("Ejemplo: 10:00  →  se re-optimiza desde las 10:00 hasta las 24:00")
    print("         El ejecutado considerado será 00:30 → 09:30")
    print()
    while True:
        hora_corte_str = input("Hora de corte (HH:MM, múltiplo de :00 o :30): ").strip()
        try:
            p = hora_a_periodo(hora_corte_str)
            if 1 <= p <= PERIODOS_TOTALES - 1:
                periodo_corte = p
                break
            else:
                print(f"  → Hora fuera de rango (01:00 – 23:30).")
        except Exception:
            print("  → Formato inválido. Use HH:MM.")

    hora_ejecutado_fin = periodo_a_hora(periodo_corte - 1)
    horas_reopt        = PERIODOS_TOTALES - periodo_corte
    print()
    print(f"  ✓ Ejecutado:        00:30 → {hora_ejecutado_fin}  ({periodo_corte} periodos)")
    print(f"  ✓ Re-optimización:  {hora_corte_str} → 24:00  ({horas_reopt} periodos)")

    # ── Volúmenes reales a la hora de corte ────────────────
    print()
    print(f"Ingrese los volúmenes REALES de los diques a las {hora_corte_str}:")
    print(f"  Rango D. Cincel:     {V_Cincel_min/1000:.0f} – {V_Cincel_max/1000:.0f} mil m³")
    print(f"  Rango D. Campanario: {V_Campanario_min/1000:.0f} – {V_Campanario_max/1000:.0f} mil m³")
    print()

    while True:
        try:
            v_c = float(input(f"  Volumen D. Cincel a las {hora_corte_str} (m³): ").strip())
            if V_Cincel_min <= v_c <= V_Cincel_max:
                V_Cincel_corte = v_c
                break
            else:
                print(f"  → Fuera del rango [{V_Cincel_min:.0f} – {V_Cincel_max:.0f}]")
        except ValueError:
            print("  → Valor numérico inválido.")

    while True:
        try:
            v_ca = float(input(f"  Volumen D. Campanario a las {hora_corte_str} (m³): ").strip())
            if V_Campanario_min <= v_ca <= V_Campanario_max:
                V_Campanario_corte = v_ca
                break
            else:
                print(f"  → Fuera del rango [{V_Campanario_min:.0f} – {V_Campanario_max:.0f}]")
        except ValueError:
            print("  → Valor numérico inválido.")

    # ── Caudal de salida del Campanario (puede variar) ─────
    print()
    print("¿El caudal de salida del Dique Campanario ha variado respecto al programa inicial?")
    resp_q = input("  (s/n): ").strip().lower()
    if resp_q == 's':
        while True:
            try:
                Q_salida_Campanario_nuevo = float(
                    input("  Nuevo caudal salida D. Campanario (m³/s): ").strip())
                if Q_salida_Campanario_nuevo > 0:
                    break
                print("  → Debe ser mayor a 0.")
            except ValueError:
                print("  → Valor numérico inválido.")

    # ── Cargar datos del programa anterior (para graficarlo) ──
    print()
    print("Para graficar el programa anterior (ejecutado / programa inicial),")
    print("ingrese el nombre del archivo Excel de resultados previos.")
    print("(Deje en blanco si no desea comparar gráficamente con el anterior)")
    archivo_anterior = input("  Nombre archivo anterior (Enter para omitir): ").strip()
    if archivo_anterior:
        try:
            df_ant = pd.read_excel(archivo_anterior, sheet_name='Resultados')
            Q_opt_anterior       = df_ant['Caudal Salida D. Cincel (m3/s)'].values[:PERIODOS_TOTALES]
            V_Cincel_anterior    = df_ant['Volumen D. Cincel (m3)'].values[:PERIODOS_TOTALES]
            V_camp_anterior      = df_ant['Volumen D. Campanario (m3)'].values[:PERIODOS_TOTALES]
            CMG_anterior         = df_ant['CMG (S/./MWh)'].values[:PERIODOS_TOTALES]
            Pot_entrada_anterior = df_ant['P_Char 5 (MW)'].values[:PERIODOS_TOTALES]
            print("  ✓ Datos del programa anterior cargados correctamente.")
        except Exception as e:
            print(f"  ✗ No se pudo cargar: {e}. Se continuará sin datos anteriores.")

    print()
    print("─" * 50)


# =============================================
# CARGA DE DATOS DE ENTRADA (REPROGRAMA O INICIAL)
# =============================================

print()
print("Cargando datos de entrada...")

df_raw = pd.read_excel("Datos_Entrada.xlsx", header=None)
Q_salida_Campanario = float(df_raw.iloc[51, 2])

# Si en reprograma se indicó un caudal nuevo, usarlo
if ES_REPROGRAMA and Q_salida_Campanario_nuevo is not None:
    print(f"  Caudal D. Campanario original : {Q_salida_Campanario:.2f} m³/s")
    Q_salida_Campanario = Q_salida_Campanario_nuevo
    print(f"  Caudal D. Campanario actualizado: {Q_salida_Campanario:.2f} m³/s")
else:
    print(f"  Caudal salida Dique Campanario: {Q_salida_Campanario:.2f} m³/s")

try:
    datos_excel = pd.read_excel("Datos_Entrada.xlsx", engine='openpyxl')
    costo_marginal_completo  = datos_excel['CMG'].values[:PERIODOS_TOTALES].astype(np.float64)
    Potencia_entrada_completa = datos_excel['P_CHAR 5'].values[:PERIODOS_TOTALES]
    Q_Cincel_completo = (Potencia_entrada_completa / 5.98).astype(np.float64)
except Exception as e:
    raise ValueError(f"Error en carga de datos: {str(e)}")

# ── Ajuste según modo ────────────────────────────────────────────────────────
if ES_REPROGRAMA:
    # Solo los periodos desde la hora de corte hasta el final
    horas = PERIODOS_TOTALES - periodo_corte
    costo_marginal   = costo_marginal_completo[periodo_corte:]
    Potencia_entrada = Potencia_entrada_completa[periodo_corte:]
    Q_Cincel         = Q_Cincel_completo[periodo_corte:]

    # Volúmenes iniciales = valores reales a la hora de corte
    V_Cincel_inicio     = V_Cincel_corte
    V_Campanario_inicio = V_Campanario_corte

    # Objetivo final: mantener 85 % del máximo al finalizar
    V_Cincel_final      = 0.85 * V_Cincel_max
    V_Campanario_final  = 0.85 * V_Campanario_max

    print(f"\n  → Optimizando {horas} periodos ({hora_corte_str} → 24:00)")
    print(f"  → V_Cincel inicial:     {V_Cincel_inicio/1000:.1f} mil m³")
    print(f"  → V_Campanario inicial: {V_Campanario_inicio/1000:.1f} mil m³")

else:
    horas = PERIODOS_TOTALES
    costo_marginal   = costo_marginal_completo
    Potencia_entrada = Potencia_entrada_completa
    Q_Cincel         = Q_Cincel_completo

    porcentaje_inicial_Cincel     = 0.85
    porcentaje_final_Cincel       = 0.85
    porcentaje_inicial_Campanario = 0.85
    porcentaje_final_Campanario   = 0.85

    V_Cincel_inicio     = porcentaje_inicial_Cincel     * V_Cincel_max
    V_Cincel_final      = porcentaje_final_Cincel       * V_Cincel_max
    V_Campanario_inicio = porcentaje_inicial_Campanario * V_Campanario_max
    V_Campanario_final  = porcentaje_final_Campanario   * V_Campanario_max


# =============================================
# FUNCIONES AUXILIARES
# =============================================

def calcular_volumenes_con_caudales(Q_salida_Cincel):
    """
    Calcula volúmenes del Dique Cincel y Dique Campanario.
    Charcani 4 y Charcani 6 toman agua al mismo tiempo (desfase < 1 intervalo).
    Devuelve: V_Cincel, V_Campanario, Q_CH4, Q_CH6
    """
    Q_salida_Cincel = np.asarray(Q_salida_Cincel, dtype=np.float64)

    V_Cincel     = np.empty(horas + 1, dtype=np.float64)
    V_Campanario = np.empty(horas + 1, dtype=np.float64)
    Q_CH4        = np.empty(horas, dtype=np.float64)
    Q_CH6        = np.empty(horas, dtype=np.float64)

    V_Cincel[0]     = V_Cincel_inicio
    V_Campanario[0] = V_Campanario_inicio

    for t in range(horas):
        V_Cincel[t+1] = V_Cincel[t] + (Q_Cincel[t] - Q_salida_Cincel[t]) * 1800.0

        if t == 0:
            Q_t = Q_salida_Cincel[0]
        else:
            Q_t = Q_salida_Cincel[t-1]

        Q_CH4[t] = Q_t
        Q_CH6[t] = Q_t

        V_Campanario[t+1] = V_Campanario[t] + (Q_t - Q_salida_Campanario) * 1800.0

    return V_Cincel, V_Campanario, Q_CH4, Q_CH6


def verificar_violaciones(V_Cincel, V_Campanario):
    """Verifica violaciones de manera detallada"""
    violaciones = {
        'Cincel_sobre_max':       np.sum(V_Cincel[1:]     > V_Cincel_max),
        'Cincel_bajo_min':        np.sum(V_Cincel[1:]     < V_Cincel_min),
        'Campanario_sobre_max':   np.sum(V_Campanario[1:] > V_Campanario_max),
        'Campanario_bajo_min':    np.sum(V_Campanario[1:] < V_Campanario_min),
        'Cincel_max_exceso':      np.maximum(V_Cincel[1:]     - V_Cincel_max,     0).sum(),
        'Cincel_min_deficit':     np.maximum(V_Cincel_min     - V_Cincel[1:],     0).sum(),
        'Campanario_max_exceso':  np.maximum(V_Campanario[1:] - V_Campanario_max, 0).sum(),
        'Campanario_min_deficit': np.maximum(V_Campanario_min - V_Campanario[1:], 0).sum(),
        'Campanario_negativo':    np.sum(V_Campanario[1:] < 0)
    }
    return violaciones


def reparar_solucion_inteligente(Q_prop):
    """
    Reparación inteligente que respeta límites.
    Combina suavizado con ajuste de límites.
    """
    Q_prop = np.asarray(Q_prop, dtype=np.float64)

    Q_suav = uniform_filter1d(Q_prop, size=3, mode='mirror')
    Q_suav = np.clip(Q_suav, Q_rango[0], Q_rango[1]).astype(np.float64)

    max_iter = 30
    for iteracion in range(max_iter):
        V_c, V_ca, _, _ = calcular_volumenes_con_caudales(Q_suav)
        viol = verificar_violaciones(V_c, V_ca)

        if (viol['Cincel_sobre_max']     == 0 and viol['Cincel_bajo_min']    == 0 and
            viol['Campanario_sobre_max'] == 0 and viol['Campanario_bajo_min'] == 0):
            break

        ajuste = np.zeros(horas, dtype=np.float64)

        for t in range(horas):
            idx = t + 1
            if idx < len(V_c):
                if V_c[idx] > V_Cincel_max * 0.98:
                    exceso = (V_c[idx] - V_Cincel_max) / 1800.0
                    ajuste[t] += min(exceso * 0.3, 1.0)
                elif V_c[idx] < V_Cincel_min * 1.02:
                    deficit = (V_Cincel_min - V_c[idx]) / 1800.0
                    ajuste[t] -= min(deficit * 0.3, 1.0)

        for t in range(1, horas + 1):
            if t < len(V_ca):
                if V_ca[t] > V_Campanario_max * 0.98:
                    exceso = (V_ca[t] - V_Campanario_max) / 1800.0
                    if t-1 >= 0:
                        ajuste[t-1] -= min(exceso * 0.3, 1.0)
                elif V_ca[t] < V_Campanario_min * 1.02:
                    deficit = (V_Campanario_min - V_ca[t]) / 1800.0
                    if t-1 >= 0:
                        ajuste[t-1] += min(deficit * 0.3, 1.0)

        factor = np.float64(0.7 * (1 - iteracion/max_iter) + 0.3)
        Q_suav = Q_suav + ajuste * factor
        Q_suav = np.clip(Q_suav, Q_rango[0], Q_rango[1])

        if iteracion % 5 == 0:
            promedio = np.mean(Q_suav)
            if abs(promedio - Q_salida_Campanario) > 0.2:
                factor_ajuste = Q_salida_Campanario / promedio
                Q_suav = Q_suav * np.float64(factor_ajuste)
                Q_suav = np.clip(Q_suav, Q_rango[0], Q_rango[1])

    V_c_final, V_ca_final, _, _ = calcular_volumenes_con_caudales(Q_suav)

    error_Cincel = V_c_final[-1] - V_Cincel_final
    if abs(error_Cincel) > 1000.0:
        ajuste_final = error_Cincel / (1800.0 * horas)
        Q_suav = Q_suav + np.float64(ajuste_final)
        Q_suav = np.clip(Q_suav, Q_rango[0], Q_rango[1])

    Q_final = uniform_filter1d(Q_suav, size=3, mode='mirror')
    Q_final = np.clip(Q_final, Q_rango[0], Q_rango[1]).astype(np.float64)

    return Q_final


# =============================================
# FUNCIÓN OBJETIVO CON PENALIDADES FUERTES
# =============================================

def funcion_objetivo_unificada(positions):
    """
    Función objetivo:
    - Penalidades fuertes por violar límites
    - Reparación inteligente de soluciones
    - Optimización de ingresos (Charcani 4 + Charcani 6)
    """
    n_particles = positions.shape[0]
    costos_totales = np.zeros(n_particles, dtype=np.float64)

    for i in range(n_particles):
        Q_reparado = reparar_solucion_inteligente(positions[i])

        V_Cincel, V_Campanario, Q_CH4, Q_CH6 = calcular_volumenes_con_caudales(Q_reparado)

        Potencia_CH4 = rendimiento_CH4 * Q_CH4
        Potencia_CH6 = rendimiento_CH6 * Q_CH6

        ingresos = 0.5 * np.sum((Potencia_CH4 + Potencia_CH6) * costo_marginal, dtype=np.float64)
        costos_base = -ingresos

        violaciones = verificar_violaciones(V_Cincel, V_Campanario)

        penal_volumenes = 0.0

        if violaciones['Cincel_sobre_max'] > 0:
            penal_volumenes += 1e8 * violaciones['Cincel_sobre_max']
            penal_volumenes += 1e6 * violaciones['Cincel_max_exceso']

        if violaciones['Cincel_bajo_min'] > 0:
            penal_volumenes += 1e8 * violaciones['Cincel_bajo_min']
            penal_volumenes += 1e6 * violaciones['Cincel_min_deficit']

        if violaciones['Campanario_sobre_max'] > 0:
            penal_volumenes += 1e8 * violaciones['Campanario_sobre_max']
            penal_volumenes += 1e6 * violaciones['Campanario_max_exceso']

        if violaciones['Campanario_bajo_min'] > 0:
            penal_volumenes += 1e8 * violaciones['Campanario_bajo_min']
            penal_volumenes += 1e6 * violaciones['Campanario_min_deficit']

        if violaciones['Campanario_negativo'] > 0:
            penal_volumenes += 1e10 * violaciones['Campanario_negativo']

        penal_final = 1e7 * float(
            abs(V_Cincel[-1]     - V_Cincel_final) +
            abs(V_Campanario[-1] - V_Campanario_final)
        )

        penal_suavidad = 1e4 * np.sum(np.square(np.diff(Q_reparado)), dtype=np.float64)

        costos_totales[i] = np.float64(costos_base + penal_volumenes +
                                        penal_final + penal_suavidad)

    return costos_totales


# =============================================
# CONFIGURACIÓN PSO
# =============================================

print("\nConfigurando PSO optimizado...")

Q_inicial_base = np.full(horas, Q_salida_Campanario, dtype=np.float64)
for t in range(horas):
    if Q_Cincel[t] > Q_salida_Campanario + 2.0:
        Q_inicial_base[t] = min(Q_inicial_base[t] + 0.8, Q_rango[1] - 0.5)
    elif Q_Cincel[t] < Q_salida_Campanario - 2.0:
        Q_inicial_base[t] = max(Q_inicial_base[t] - 0.8, Q_rango[0] + 0.5)

n_particles = 150
init_pos = np.tile(Q_inicial_base, (n_particles, 1))
for i in range(n_particles):
    noise = np.random.randn(horas) * 0.4
    init_pos[i] = np.clip(init_pos[i] + noise, Q_rango[0], Q_rango[1])

init_pos = init_pos.astype(np.float64)

options = {
    'c1': 2.0,
    'c2': 2.0,
    'w': 0.9,
    'v_max': 1.5
}

optimizador = ps.single.GlobalBestPSO(
    n_particles=n_particles,
    dimensions=horas,
    options=options,
    bounds=(np.array([Q_rango[0]]*horas, dtype=np.float64),
            np.array([Q_rango[1]]*horas, dtype=np.float64)),
    init_pos=init_pos
)

# =============================================
# OPTIMIZACIÓN CON INERCIA DINÁMICA (DECAY)
# =============================================

print("Iniciando optimización con inercia dinámica...")
print(f"Configuración: {n_particles} partículas × 150 iteraciones")
if ES_REPROGRAMA:
    print(f"Modo REPROGRAMA: {hora_corte_str} → 24:00 ({horas} periodos)")
else:
    print("Modo PROGRAMA INICIAL: 00:30 → 24:00 (48 periodos)")
print(f"Inercia inicial: 0.9 | Inercia final: 0.4")

tiempo_inicio_opt = time.time()

w_inicial = 0.9
w_final   = 0.4
max_iter  = 150

cost_history = []
w_history    = []

best_cost, best_pos = None, None

for iteracion in range(max_iter):
    w_actual = w_final + (w_inicial - w_final) * (1 - iteracion / max_iter)
    optimizador.options['w'] = np.float64(w_actual)

    try:
        cost, pos = optimizador.optimize(funcion_objetivo_unificada, iters=1, verbose=False)
        cost = np.float64(cost)
        pos  = np.asarray(pos, dtype=np.float64)
    except Exception as e:
        print(f"Error en iteración {iteracion + 1}: {e}")
        if best_pos is not None:
            pos  = best_pos.copy()
            cost = best_cost
        else:
            pos  = init_pos[0].copy()
            cost = np.float64(1e15)

    cost_history.append(cost)
    w_history.append(w_actual)

    if (iteracion + 1) % 50 == 0:
        print(f"Iteración {iteracion + 1}/{max_iter} - Costo: {cost:.2e} - Inercia: {w_actual:.3f}")

    if best_cost is None or cost < best_cost:
        best_cost = cost
        best_pos  = pos.copy()

tiempo_opt = time.time() - tiempo_inicio_opt

print(f"\nOptimización completada en {tiempo_opt/60:.2f} minutos")
print(f"Mejor costo encontrado: {best_cost:.2e}")

# Reparación final
Q_opt = reparar_solucion_inteligente(best_pos)

# =============================================
# CÁLCULO FINAL Y VERIFICACIÓN
# =============================================

print("\n" + "="*70)
print("VERIFICACIÓN FINAL DE LÍMITES")
print("="*70)

V_Cincel, V_Campanario, Q_CH4_opt, Q_CH6_opt = calcular_volumenes_con_caudales(Q_opt)

Potencia_CH4 = rendimiento_CH4 * Q_CH4_opt
Potencia_CH6 = rendimiento_CH6 * Q_CH6_opt

Costos_horarios = 0.5 * (Potencia_CH4 + Potencia_CH6) * costo_marginal
Ingresos = np.sum(Costos_horarios)

violaciones = verificar_violaciones(V_Cincel, V_Campanario)

print(f"\n1. ESTADO DE LÍMITES:")
print(f"   Dique Cincel:")
print(f"     Sobre máximo: {violaciones['Cincel_sobre_max']} periodos")
print(f"     Bajo mínimo:  {violaciones['Cincel_bajo_min']} periodos")
print(f"   Dique Campanario:")
print(f"     Sobre máximo: {violaciones['Campanario_sobre_max']} periodos")
print(f"     Bajo mínimo:  {violaciones['Campanario_bajo_min']} periodos")

print(f"\n2. CONDICIONES FINALES:")
print(f"   D. Cincel:     {V_Cincel[-1]/1000:.1f} vs {V_Cincel_final/1000:.1f} mil m³")
print(f"   D. Campanario: {V_Campanario[-1]/1000:.1f} vs {V_Campanario_final/1000:.1f} mil m³")

print(f"\n3. CAUDALES OPTIMIZADOS:")
print(f"   Mínimo:  {np.min(Q_opt):.2f} m³/s")
print(f"   Máximo:  {np.max(Q_opt):.2f} m³/s")
print(f"   Promedio: {np.mean(Q_opt):.2f} m³/s (objetivo: {Q_salida_Campanario:.2f})")

correlacion = np.corrcoef(costo_marginal, Potencia_CH4 + Potencia_CH6)[0, 1]
print(f"\n4. DESPACHO ECONÓMICO:")
print(f"   Correlación CMG-Potencia total: {correlacion:.3f}")
print(f"   Ingresos totales (período optimizado): ${Ingresos:,.2f}")

es_valida = (
    violaciones['Cincel_sobre_max']     == 0 and
    violaciones['Cincel_bajo_min']      == 0 and
    violaciones['Campanario_sobre_max'] == 0 and
    violaciones['Campanario_bajo_min']  == 0 and
    violaciones['Campanario_negativo']  == 0 and
    abs(V_Cincel[-1]     - V_Cincel_final)     < 5000 and
    abs(V_Campanario[-1] - V_Campanario_final) < 5000
)

print("\n" + "="*70)
if es_valida:
    print("✓ SOLUCIÓN VÁLIDA - TODOS LOS LÍMITES RESPETADOS")
else:
    print("✗ SOLUCIÓN NO VÁLIDA - REVISAR RESTRICCIONES")
    if violaciones['Cincel_sobre_max'] > 0:
        print(f"   • Dique Cincel excede máximo en {violaciones['Cincel_sobre_max']} periodos")
    if violaciones['Campanario_negativo'] > 0:
        print(f"   • Dique Campanario tiene {violaciones['Campanario_negativo']} periodos negativos")
print("="*70)


# =============================================
# ETIQUETAS DE TIEMPO PARA GRÁFICOS
# =============================================

def generar_etiquetas(periodo_inicio, n_periodos):
    """Genera etiquetas HH:MM para cada periodo. periodo_inicio es 0-based."""
    etiquetas = []
    for i in range(n_periodos):
        p = periodo_inicio + i + 1    # +1 porque los periodos se etiquetan al final
        minutos = p * 30
        h = (minutos // 60) % 24
        m = minutos % 60
        if h == 0 and m == 0:
            etiquetas.append("24:00")
        else:
            etiquetas.append(f"{h:02d}:{m:02d}")
    return etiquetas

# Eje X del periodo re-optimizado (o completo si es programa inicial)
periodos_reopt = np.arange(periodo_corte + 1, periodo_corte + horas + 1)  # valores 1-based globales
etiquetas_reopt = generar_etiquetas(periodo_corte, horas)

# Titulo de modo
titulo_modo = f"REPROGRAMA desde {hora_corte_str}" if ES_REPROGRAMA else "PROGRAMA INICIAL"

# Color del período re-optimizado y del ejecutado
COLOR_REOPT   = 'tab:blue'
COLOR_EJEC    = 'silver'
COLOR_REOPT2  = 'tab:green'
COLOR_CMG_REOPT = 'tab:red'
COLOR_CMG_EJEC  = 'lightcoral'


# =============================================
# GRÁFICOS
# =============================================

print("\nGenerando gráficos...")


# ──────────────────────────────────────────────
# 1. GRÁFICO DE CAUDALES
# ──────────────────────────────────────────────
fig1, ax1 = plt.subplots(figsize=(10, 5))

if ES_REPROGRAMA and Q_opt_anterior is not None:
    # Parte ejecutada (periodos 0..periodo_corte-1, índice global)
    per_ejec = np.arange(1, periodo_corte + 1)
    ax1.plot(per_ejec, Q_opt_anterior[:periodo_corte],
             color=COLOR_EJEC, linewidth=2, label='Programa anterior (ejecutado)', zorder=2)

# Período re-optimizado
ax1.plot(periodos_reopt, Q_opt, color=COLOR_REOPT, linewidth=2.2,
         label=f'Q optimizado — {titulo_modo}', zorder=3)
ax1.plot(periodos_reopt, Q_Cincel, color='darkgreen', linestyle='--',
         linewidth=1, alpha=0.5, label='Q entrada D. Cincel (reprograma)', zorder=2)
ax1.axhline(y=np.mean(Q_opt), color='r', linestyle='--', alpha=0.5,
            label=f'Promedio reopt: {np.mean(Q_opt):.1f} m³/s')
ax1.fill_between(periodos_reopt, Q_rango[0], Q_rango[1],
                 color='gray', alpha=0.15, label='Límites operativos')

if ES_REPROGRAMA:
    ax1.axvline(x=periodo_corte + 0.5, color='orange', linewidth=2,
                linestyle='--', label=f'Hora de corte: {hora_corte_str}', zorder=4)

ax1.set_title(f'Caudal Salida D. Cincel — {titulo_modo}  (Q_Camp={Q_salida_Campanario:.1f} m³/s)')
ax1.set_xlabel('Periodo (30 min)')
ax1.set_ylabel('Caudal (m³/s)')
ax1.legend(loc='upper left', fontsize=9)
ax1.grid(True, alpha=0.3)

# Etiquetas del eje X: mostrar cada 4 periodos aprox
todos_periodos = np.arange(1, PERIODOS_TOTALES + 1)
todas_etiquetas = generar_etiquetas(0, PERIODOS_TOTALES)
tick_step = 4
ax1.set_xticks(todos_periodos[::tick_step])
ax1.set_xticklabels(todas_etiquetas[::tick_step], rotation=45, fontsize=7)
ax1.set_xlim(0.5, PERIODOS_TOTALES + 0.5)

fig1.tight_layout()
fig1.savefig('grafico_caudal.png', dpi=150, bbox_inches='tight')


# ──────────────────────────────────────────────
# 2. GRÁFICO DE DESPACHO ECONÓMICO (CMG + Potencias)
# ──────────────────────────────────────────────
fig2, ax2 = plt.subplots(figsize=(10, 5))
ax2_b = ax2.twinx()

if ES_REPROGRAMA and Q_opt_anterior is not None:
    # CMG y potencias del ejecutado (en gris)
    per_ejec = np.arange(1, periodo_corte + 1)
    ax2.plot(per_ejec, CMG_anterior[:periodo_corte],
             color=COLOR_CMG_EJEC, linewidth=1.5, alpha=0.7, label='CMG anterior')
    P_CH4_ant = rendimiento_CH4 * Q_opt_anterior[:periodo_corte]
    P_CH6_ant = rendimiento_CH6 * Q_opt_anterior[:periodo_corte]
    ax2_b.plot(per_ejec, P_CH4_ant, color='lightblue',  linewidth=1.5,
               linestyle='--', alpha=0.7, label='CH4 anterior')
    ax2_b.plot(per_ejec, P_CH6_ant, color='lightgreen', linewidth=1.5,
               linestyle='--', alpha=0.7, label='CH6 anterior')

# Re-optimizado
ax2.plot(periodos_reopt, costo_marginal, color=COLOR_CMG_REOPT, linewidth=2,
         label='CMG (reprograma)', alpha=0.9)
ax2.fill_between(periodos_reopt, 0, costo_marginal, color='red', alpha=0.08)
ax2_b.plot(periodos_reopt, Potencia_CH4, color='tab:blue',  linewidth=2,
           label='Charcani 4 (reopt)', alpha=0.9)
ax2_b.plot(periodos_reopt, Potencia_CH6, color='tab:green', linewidth=2,
           linestyle='--', label='Charcani 6 (reopt)', alpha=0.9)

if ES_REPROGRAMA:
    ax2.axvline(x=periodo_corte + 0.5, color='orange', linewidth=2,
                linestyle='--', label=f'Hora de corte: {hora_corte_str}', zorder=4)

ax2.set_title(f'Despacho Económico — {titulo_modo}  (Corr: {correlacion:.3f})')
ax2.set_xlabel('Periodo (30 min)')
ax2.set_ylabel('CMG (S/./MWh)', color='r')
ax2_b.set_ylabel('Potencia (MW)', color='b')
ax2.tick_params(axis='y', labelcolor='r')
ax2_b.tick_params(axis='y', labelcolor='b')
ax2.legend(loc='upper left', fontsize=8)
ax2_b.legend(loc='upper right', fontsize=8)
ax2.grid(True, alpha=0.3)
ax2.set_xticks(todos_periodos[::tick_step])
ax2.set_xticklabels(todas_etiquetas[::tick_step], rotation=45, fontsize=7)
ax2.set_xlim(0.5, PERIODOS_TOTALES + 0.5)

fig2.tight_layout()
fig2.savefig('grafico_cmg_potencia.png', dpi=150, bbox_inches='tight')


# ──────────────────────────────────────────────
# 3. GRÁFICO DE VOLÚMENES
# ──────────────────────────────────────────────
plt.rcParams['axes.grid']      = True
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['axes.labelsize'] = 10

fig3, ax3 = plt.subplots(figsize=(10, 5))

if ES_REPROGRAMA and V_Cincel_anterior is not None:
    per_ejec = np.arange(1, periodo_corte + 1)
    ax3.plot(per_ejec, V_Cincel_anterior[:periodo_corte],
             color='silver', linewidth=2, linestyle='-', label='Vol. Cincel anterior')
    ax3.plot(per_ejec, V_camp_anterior[:periodo_corte],
             color='lightgreen', linewidth=2, linestyle='-', label='Vol. Campanario anterior')

# Re-optimizado
ax3.plot(periodos_reopt, V_Cincel[1:],     color='tab:blue',  linewidth=2.2,
         label='Vol. D. Cincel (reopt)')
ax3.plot(periodos_reopt, V_Campanario[1:], color='tab:green', linewidth=2.2,
         label='Vol. D. Campanario (reopt)')
ax3.fill_between(periodos_reopt, V_Cincel_min,     V_Cincel_max,
                 color='blue',  alpha=0.07)
ax3.fill_between(periodos_reopt, V_Campanario_min, V_Campanario_max,
                 color='green', alpha=0.07)

# Líneas de límites
ax3.axhline(y=V_Cincel_max,     color='navy',  linestyle=':', linewidth=1, alpha=0.5)
ax3.axhline(y=V_Cincel_min,     color='navy',  linestyle=':', linewidth=1, alpha=0.5)
ax3.axhline(y=V_Campanario_max, color='darkgreen', linestyle=':', linewidth=1, alpha=0.5)
ax3.axhline(y=V_Campanario_min, color='darkgreen', linestyle=':', linewidth=1, alpha=0.5)

# Marcador del volumen real en la hora de corte
if ES_REPROGRAMA:
    ax3.axvline(x=periodo_corte + 0.5, color='orange', linewidth=2,
                linestyle='--', label=f'Hora de corte: {hora_corte_str}', zorder=4)
    ax3.plot(periodo_corte + 1, V_Cincel_inicio, 'o',
             color='tab:blue', markersize=9, zorder=5,
             label=f'V_Cincel real: {V_Cincel_inicio/1000:.1f} mil m³')
    ax3.plot(periodo_corte + 1, V_Campanario_inicio, 's',
             color='tab:green', markersize=9, zorder=5,
             label=f'V_Camp real: {V_Campanario_inicio/1000:.1f} mil m³')

# Violaciones
if violaciones['Cincel_sobre_max'] > 0:
    viol_idx = np.where(V_Cincel[1:] > V_Cincel_max)[0]
    ax3.plot(periodos_reopt[viol_idx], V_Cincel[1:][viol_idx],
             'rX', markersize=10, label='Violación máximo')
if violaciones['Cincel_bajo_min'] > 0:
    viol_idx = np.where(V_Cincel[1:] < V_Cincel_min)[0]
    ax3.plot(periodos_reopt[viol_idx], V_Cincel[1:][viol_idx],
             'rX', markersize=10, label='Violación mínimo')
if violaciones['Campanario_sobre_max'] > 0:
    viol_idx = np.where(V_Campanario[1:] > V_Campanario_max)[0]
    ax3.plot(periodos_reopt[viol_idx], V_Campanario[1:][viol_idx], 'rX', markersize=10)
if violaciones['Campanario_bajo_min'] > 0:
    viol_idx = np.where(V_Campanario[1:] < V_Campanario_min)[0]
    ax3.plot(periodos_reopt[viol_idx], V_Campanario[1:][viol_idx], 'rX', markersize=10)
if violaciones['Campanario_negativo'] > 0:
    viol_idx = np.where(V_Campanario[1:] < 0)[0]
    ax3.plot(periodos_reopt[viol_idx], V_Campanario[1:][viol_idx],
             'kX', markersize=10, label='Volumen negativo')

ax3.set_title(f'Evolución de Volúmenes — {titulo_modo}')
ax3.set_xlabel('Periodo (30 minutos)')
ax3.set_ylabel('Volumen (m³)')
ax3.legend(loc='upper left', fontsize=8)
ax3.set_xticks(todos_periodos[::tick_step])
ax3.set_xticklabels(todas_etiquetas[::tick_step], rotation=45, fontsize=7)
ax3.set_xlim(0.5, PERIODOS_TOTALES + 0.5)

fig3.tight_layout()
fig3.savefig('grafico_volumenes.png', dpi=150, bbox_inches='tight')

print("✓ Gráficos generados\n")


# =============================================
# GUARDADO DE RESULTADOS EN EXCEL
# =============================================

print("Guardando resultados en Excel...")

# Etiquetas de hora para los periodos re-optimizados
horas_etiquetas_reopt = etiquetas_reopt

# Columna CH 123
CH123 = []
for t in range(horas):
    CH123.append(6.6 if Q_salida_Campanario >= 10.0 else None)

sufijo = f"_REPROG_{hora_corte_str.replace(':', 'h')}" if ES_REPROGRAMA else ""
nombre_archivo = f"Resultados_PSO_CH46_Q{int(Q_salida_Campanario)}{sufijo}.xlsx"

resultados = pd.DataFrame({
    'HORA':                               horas_etiquetas_reopt,
    'P_Char 5 (MW)':                      np.round(Potencia_entrada.astype(float), 1),
    'Volumen D. Cincel (m3)':             V_Cincel[1:].astype(int),
    'Caudal Salida D. Cincel (m3/s)':     np.round(Q_opt, 1),
    'Volumen D. Campanario (m3)':         V_Campanario[1:].astype(int),
    'Caudal salida D. Campanario (m3/s)': np.round(np.full(horas, Q_salida_Campanario), 1),
    'CH 4 (MW)':                          np.round(Potencia_CH4, 1),
    'CH 6 (MW)':                          np.round(Potencia_CH6, 1),
    'CH 123 (MW)':                        CH123,
    'CMG (S/./MWh)':                      np.round(costo_marginal, 1),
    'Ingresos':                           np.round(Costos_horarios, 2),
})

with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
    resultados.to_excel(writer, sheet_name='Resultados', index=False)

# --- Formato con openpyxl ---
wb = load_workbook(nombre_archivo)
ws = wb['Resultados']

header_font  = Font(bold=True, color="FFFFFF")
header_fill  = PatternFill("solid", fgColor="1F497D")
center_align = Alignment(horizontal='center', vertical='center')
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

for col in range(1, 12):
    cell = ws.cell(row=1, column=col)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = thin_border

anchos = [8, 14, 22, 28, 26, 32, 10, 10, 12, 16, 12]
for i, ancho in enumerate(anchos, start=1):
    ws.column_dimensions[get_column_letter(i)].width = ancho

for row in range(2, horas + 2):
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        cell.alignment = center_align
        cell.border    = thin_border

# Resumen
fila_resumen = horas + 3
resumen_datos = [
    ("Modo",                            titulo_modo),
    ("Ingresos totales período reopt ($)", round(Ingresos, 2)),
    ("Tiempo optimización (min)",        round(tiempo_opt/60, 3)),
    ("Tiempo total (min)",               round((time.time()-start_time)/60, 3)),
    ("Correlación CMG-Potencia total",   round(correlacion, 4)),
    ("Q promedio optimizado (m³/s)",     round(float(np.mean(Q_opt)), 3)),
    ("Vol. D. Cincel inicial (mil m³)",  round(V_Cincel_inicio/1000, 1)),
    ("Vol. D. Cincel final (mil m³)",    round(float(V_Cincel[-1])/1000, 1)),
    ("Vol. D. Campanario inicial (mil m³)", round(V_Campanario_inicio/1000, 1)),
    ("Vol. D. Campanario final (mil m³)",   round(float(V_Campanario[-1])/1000, 1)),
    ("Solución válida",                  'SÍ' if es_valida else 'NO - REVISAR'),
]
if ES_REPROGRAMA:
    resumen_datos.insert(1, ("Hora de corte", hora_corte_str))

for idx, (etiq, val) in enumerate(resumen_datos):
    ws.cell(row=fila_resumen + idx, column=1, value=etiq)
    ws.cell(row=fila_resumen + idx, column=2, value=val)

# Gráficos
imagenes = [
    ('grafico_caudal.png',       'M2'),
    ('grafico_cmg_potencia.png', 'M25'),
    ('grafico_volumenes.png',    'M48'),
]
for img_file, cell_anchor in imagenes:
    try:
        img = Image(img_file)
        img.width  = int(img.width  * 0.55)
        img.height = int(img.height * 0.55)
        img.anchor = cell_anchor
        ws.add_image(img)
    except Exception as e:
        print(f"Advertencia: No se pudo insertar {img_file}: {e}")

wb.save(nombre_archivo)
print(f"✓ Excel guardado: {nombre_archivo}")


# =============================================
# RESUMEN FINAL
# =============================================

tiempo_total = (time.time() - start_time) / 60
print(f"\n{'='*70}")
print(f"OPTIMIZACIÓN PSO — CHARCANI 4 & 6 — {titulo_modo.upper()} — COMPLETADA")
if ES_REPROGRAMA:
    print(f"Período re-optimizado: {hora_corte_str} → 24:00  ({horas} periodos de 30 min)")
    print(f"Vol. Cincel en hora de corte:     {V_Cincel_inicio/1000:.1f} mil m³")
    print(f"Vol. Campanario en hora de corte: {V_Campanario_inicio/1000:.1f} mil m³")
print(f"Ingresos período optimizado: ${Ingresos:,.2f}")
print(f"Potencia promedio CH4: {np.mean(Potencia_CH4):.2f} MW  |  CH6: {np.mean(Potencia_CH6):.2f} MW")
print(f"Tiempo total: {tiempo_total:.2f} min")
print(f"{'='*70}")
