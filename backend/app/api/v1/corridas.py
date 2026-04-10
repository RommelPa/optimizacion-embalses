from typing import Optional
from uuid import uuid4

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel, Field

from app.integrations.pso.contracts import PSOWrapperInput
from app.integrations.pso.wrapper import ejecutar_corrida_pso
from app.integrations.pso.errors import PSOExecutionError, PSOValidationError

import json

from app.db.session import SessionLocal
from app.models.corrida import Corrida

import csv
import io

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from datetime import timezone

def serialize_datetime_utc(dt):
    if dt is None:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

router = APIRouter()


class CorridaCreateRequest(BaseModel):
    modo_operacion: str = Field(..., examples=["inicial"])
    fecha_proceso: str = Field(..., examples=["2026-04-08"])
    escenario: str = Field(..., examples=["base"])
    origen_datos: str = Field(..., examples=["manual"])
    observaciones: Optional[str] = Field(
        default=None,
        examples=["Primera corrida refinada"],
    )
    archivo_entrada: Optional[str] = Field(
        default=None,
        examples=["../data_samples/Datos_Entrada.xlsx"],
    )


class CorridaDataResponse(BaseModel):
    id: str
    estado: str
    modo_operacion: str
    fecha_proceso: str
    escenario: str
    origen_datos: str
    observaciones: Optional[str] = None
    version_modelo: str
    modo_ejecucion: str
    mensaje_modelo: str
    best_cost: float
    execution_time_sec: float
    q_opt: list[float]


class CorridaCreateResponse(BaseModel):
    status: str
    message: str
    data: CorridaDataResponse

class CorridaListItemResponse(BaseModel):
    id: str
    created_at: str
    fecha_proceso: str
    modo_operacion: str
    escenario: str
    origen_datos: str
    estado: str
    version_modelo: str
    modo_ejecucion: str
    best_cost: float
    execution_time_sec: float

class CorridaListResponse(BaseModel):
    items: list[CorridaListItemResponse]
    total: int

class CorridaDetailResponse(BaseModel):
    id: str
    created_at: str
    fecha_proceso: str
    modo_operacion: str
    escenario: str
    origen_datos: str
    observaciones: Optional[str] = None

    estado: str
    version_modelo: str
    modo_ejecucion: str
    mensaje_modelo: str

    best_cost: float
    execution_time_sec: float
    q_opt: list[float]
    v_cincel: list[float]
    v_campanario: list[float]
    cmg: list[float]
    potencia_ch4: list[float]
    potencia_ch6: list[float]
    ingreso: list[float]

    input_payload_json: str
    error_message: Optional[str] = None

@router.post("/corridas", response_model=CorridaCreateResponse)
def crear_corrida(payload: CorridaCreateRequest):
    corrida_id = str(uuid4())

    wrapper_input = PSOWrapperInput(
        corrida_id=corrida_id,
        modo_operacion=payload.modo_operacion,
        fecha_proceso=payload.fecha_proceso,
        escenario=payload.escenario,
        origen_datos=payload.origen_datos,
        observaciones=payload.observaciones,
        archivo_entrada=payload.archivo_entrada,
    )

    input_payload_json = json.dumps(payload.model_dump())

    db = SessionLocal()
    try:
        try:
            wrapper_result = ejecutar_corrida_pso(wrapper_input)

            corrida_db = Corrida(
                id=corrida_id,
                fecha_proceso=payload.fecha_proceso,
                modo_operacion=payload.modo_operacion,
                escenario=payload.escenario,
                origen_datos=payload.origen_datos,
                observaciones=payload.observaciones,
                estado=wrapper_result.estado,
                version_modelo=wrapper_result.version_modelo,
                modo_ejecucion=wrapper_result.modo_ejecucion,
                mensaje_modelo=wrapper_result.mensaje_modelo,
                best_cost=wrapper_result.best_cost,
                execution_time_sec=wrapper_result.execution_time_sec,
                q_opt_json=json.dumps(wrapper_result.q_opt),
                v_cincel_json=json.dumps(wrapper_result.v_cincel),
                v_campanario_json=json.dumps(wrapper_result.v_campanario),
                cmg_json=json.dumps(wrapper_result.cmg),
                potencia_ch4_json=json.dumps(wrapper_result.potencia_ch4),
                potencia_ch6_json=json.dumps(wrapper_result.potencia_ch6),
                ingreso_json=json.dumps(wrapper_result.ingreso),
                input_payload_json=input_payload_json,
                error_message=None,
            )
            db.add(corrida_db)
            db.commit()

            return CorridaCreateResponse(
                status="accepted",
                message="Corrida registrada correctamente",
                data=CorridaDataResponse(
                    id=corrida_id,
                    estado=wrapper_result.estado,
                    modo_operacion=payload.modo_operacion,
                    fecha_proceso=payload.fecha_proceso,
                    escenario=payload.escenario,
                    origen_datos=payload.origen_datos,
                    observaciones=payload.observaciones,
                    version_modelo=wrapper_result.version_modelo,
                    modo_ejecucion=wrapper_result.modo_ejecucion,
                    mensaje_modelo=wrapper_result.mensaje_modelo,
                    best_cost=wrapper_result.best_cost,
                    execution_time_sec=wrapper_result.execution_time_sec,
                    q_opt=wrapper_result.q_opt,
                ),
            )

        except PSOValidationError as exc:
            corrida_db = Corrida(
                id=corrida_id,
                fecha_proceso=payload.fecha_proceso,
                modo_operacion=payload.modo_operacion,
                escenario=payload.escenario,
                origen_datos=payload.origen_datos,
                observaciones=payload.observaciones,
                estado="fallida",
                version_modelo="pso-engine-v1",
                modo_ejecucion="error_validacion",
                mensaje_modelo="La corrida falló por error de validación",
                best_cost=0.0,
                execution_time_sec=0.0,
                q_opt_json="[]",
                v_cincel_json="[]",
                v_campanario_json="[]",
                cmg_json="[]",
                potencia_ch4_json="[]",
                potencia_ch6_json="[]",
                ingreso_json="[]",
                input_payload_json=input_payload_json,
                error_message=str(exc),
            )
            db.add(corrida_db)
            db.commit()
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        except PSOExecutionError as exc:
            corrida_db = Corrida(
                id=corrida_id,
                fecha_proceso=payload.fecha_proceso,
                modo_operacion=payload.modo_operacion,
                escenario=payload.escenario,
                origen_datos=payload.origen_datos,
                observaciones=payload.observaciones,
                estado="fallida",
                version_modelo="pso-engine-v1",
                modo_ejecucion="error_ejecucion",
                mensaje_modelo="La corrida falló durante la ejecución del motor",
                best_cost=0.0,
                execution_time_sec=0.0,
                q_opt_json="[]",
                v_cincel_json="[]",
                v_campanario_json="[]",
                cmg_json="[]",
                potencia_ch4_json="[]",
                potencia_ch6_json="[]",
                ingreso_json="[]",
                input_payload_json=input_payload_json,
                error_message=str(exc),
            )
            db.add(corrida_db)
            db.commit()
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    finally:
        db.close()

@router.get("/corridas", response_model=CorridaListResponse)
def listar_corridas(
    origen_datos: Optional[str] = Query(default=None),
    estado: Optional[str] = Query(default=None),
    id_contains: Optional[str] = Query(default=None),
    fecha_proceso: Optional[str] = Query(default=None),
):
    db = SessionLocal()
    try:
        query = db.query(Corrida)

        if origen_datos:
            query = query.filter(Corrida.origen_datos == origen_datos)

        if estado:
            query = query.filter(Corrida.estado == estado)

        if fecha_proceso:
            query = query.filter(Corrida.fecha_proceso == fecha_proceso)

        if id_contains:
            query = query.filter(Corrida.id.contains(id_contains))

        rows = query.order_by(Corrida.created_at.desc()).all()

        items = [
            CorridaListItemResponse(
                id=row.id,
                created_at=serialize_datetime_utc(row.created_at),
                fecha_proceso=row.fecha_proceso,
                modo_operacion=row.modo_operacion,
                escenario=row.escenario,
                origen_datos=row.origen_datos,
                estado=row.estado,
                version_modelo=row.version_modelo,
                modo_ejecucion=row.modo_ejecucion,
                best_cost=row.best_cost,
                execution_time_sec=row.execution_time_sec,
            )
            for row in rows
        ]

        return CorridaListResponse(
            items=items,
            total=len(items),
        )
    finally:
        db.close()

@router.get("/corridas/{corrida_id}", response_model=CorridaDetailResponse)
def obtener_corrida(corrida_id: str):
    db = SessionLocal()
    try:
        row = db.query(Corrida).filter(Corrida.id == corrida_id).first()

        if not row:
            raise HTTPException(status_code=404, detail="Corrida no encontrada")

        return CorridaDetailResponse(
            id=row.id,
            created_at=serialize_datetime_utc(row.created_at),
            fecha_proceso=row.fecha_proceso,
            modo_operacion=row.modo_operacion,
            escenario=row.escenario,
            origen_datos=row.origen_datos,
            observaciones=row.observaciones,
            estado=row.estado,
            version_modelo=row.version_modelo,
            modo_ejecucion=row.modo_ejecucion,
            mensaje_modelo=row.mensaje_modelo,
            best_cost=row.best_cost,
            execution_time_sec=row.execution_time_sec,
            q_opt=json.loads(row.q_opt_json),
            v_cincel=json.loads(row.v_cincel_json),
            v_campanario=json.loads(row.v_campanario_json),
            cmg=json.loads(row.cmg_json),
            potencia_ch4=json.loads(row.potencia_ch4_json),
            potencia_ch6=json.loads(row.potencia_ch6_json),
            ingreso=json.loads(row.ingreso_json),
            input_payload_json=row.input_payload_json,
            error_message=row.error_message,
        )
    finally:
        db.close()

@router.get("/corridas/{corrida_id}/export")
def exportar_corrida_json(corrida_id: str):
    db = SessionLocal()
    try:
        row = db.query(Corrida).filter(Corrida.id == corrida_id).first()

        if not row:
            raise HTTPException(status_code=404, detail="Corrida no encontrada")

        payload = {
            "id": row.id,
            "created_at": serialize_datetime_utc(row.created_at),
            "fecha_proceso": row.fecha_proceso,
            "modo_operacion": row.modo_operacion,
            "escenario": row.escenario,
            "origen_datos": row.origen_datos,
            "observaciones": row.observaciones,
            "estado": row.estado,
            "version_modelo": row.version_modelo,
            "modo_ejecucion": row.modo_ejecucion,
            "mensaje_modelo": row.mensaje_modelo,
            "best_cost": row.best_cost,
            "execution_time_sec": row.execution_time_sec,
            "q_opt": json.loads(row.q_opt_json),
            "v_cincel": json.loads(row.v_cincel_json),
            "v_campanario": json.loads(row.v_campanario_json),
            "cmg": json.loads(row.cmg_json),
            "potencia_ch4": json.loads(row.potencia_ch4_json),
            "potencia_ch6": json.loads(row.potencia_ch6_json),
            "ingreso": json.loads(row.ingreso_json),
            "input_payload_json": row.input_payload_json,
            "error_message": row.error_message,
        }

        content = json.dumps(payload, ensure_ascii=False, indent=2)

        return Response(
            content=content,
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="corrida_{row.id}.json"'
            },
        )
    finally:
        db.close()

@router.get("/corridas/{corrida_id}/export/csv")
def exportar_corrida_csv(corrida_id: str):
    db = SessionLocal()
    try:
        row = db.query(Corrida).filter(Corrida.id == corrida_id).first()

        if not row:
            raise HTTPException(status_code=404, detail="Corrida no encontrada")

        q_opt = json.loads(row.q_opt_json)
        v_cincel = json.loads(row.v_cincel_json)
        v_campanario = json.loads(row.v_campanario_json)
        cmg = json.loads(row.cmg_json)
        potencia_ch4 = json.loads(row.potencia_ch4_json)
        potencia_ch6 = json.loads(row.potencia_ch6_json)

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(
            [
                "periodo",
                "q_opt",
                "v_cincel",
                "v_campanario",
                "cmg",
                "potencia_ch4",
                "potencia_ch6",
            ]
        )

        n = len(q_opt)

        for i in range(n):
            writer.writerow(
                [
                    i + 1,
                    q_opt[i] if i < len(q_opt) else "",
                    v_cincel[i + 1] if i + 1 < len(v_cincel) else "",
                    v_campanario[i + 1] if i + 1 < len(v_campanario) else "",
                    cmg[i] if i < len(cmg) else "",
                    potencia_ch4[i] if i < len(potencia_ch4) else "",
                    potencia_ch6[i] if i < len(potencia_ch6) else "",
                ]
            )

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="corrida_{row.id}.csv"'
            },
        )
    finally:
        db.close()

@router.get("/corridas/{corrida_id}/export/xlsx")
def exportar_corrida_excel(corrida_id: str):
    db = SessionLocal()
    try:
        corrida = db.query(Corrida).filter(Corrida.id == corrida_id).first()

        if not corrida:
            raise HTTPException(status_code=404, detail="Corrida no encontrada")

        q_opt = json.loads(corrida.q_opt_json)
        v_cincel = json.loads(corrida.v_cincel_json)
        v_campanario = json.loads(corrida.v_campanario_json)
        cmg = json.loads(corrida.cmg_json)
        potencia_ch4 = json.loads(corrida.potencia_ch4_json)
        potencia_ch6 = json.loads(corrida.potencia_ch6_json)

        ingresos = [
            0.5
            * (
                (potencia_ch4[i] if i < len(potencia_ch4) else 0)
                + (potencia_ch6[i] if i < len(potencia_ch6) else 0)
            )
            * (cmg[i] if i < len(cmg) else 0)
            for i in range(len(q_opt))
        ]
        ingreso_total_estimado = sum(ingresos)

        wb = Workbook()

        # Hoja Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        metadata_rows = [
            ("id", corrida.id),
            ("created_at", serialize_datetime_utc(corrida.created_at)),
            ("fecha_proceso", corrida.fecha_proceso),
            ("modo_operacion", corrida.modo_operacion),
            ("escenario", corrida.escenario),
            ("origen_datos", corrida.origen_datos),
            ("observaciones", corrida.observaciones or ""),
            ("estado", corrida.estado),
            ("version_modelo", corrida.version_modelo),
            ("modo_ejecucion", corrida.modo_ejecucion),
            ("mensaje_modelo", corrida.mensaje_modelo),
            ("best_cost", corrida.best_cost),
            ("execution_time_sec", corrida.execution_time_sec),
            ("error_message", corrida.error_message or ""),
        ]

        metric_rows = [
            ("q_opt_periodos", len(q_opt)),
            ("q_opt_promedio", sum(q_opt) / len(q_opt) if q_opt else 0),
            ("q_opt_minimo", min(q_opt) if q_opt else 0),
            ("q_opt_maximo", max(q_opt) if q_opt else 0),
            ("v_cincel_puntos", len(v_cincel)),
            ("v_cincel_inicial", v_cincel[0] if v_cincel else 0),
            ("v_cincel_final", v_cincel[-1] if v_cincel else 0),
            ("v_cincel_minimo", min(v_cincel) if v_cincel else 0),
            ("v_cincel_maximo", max(v_cincel) if v_cincel else 0),
            ("v_campanario_puntos", len(v_campanario)),
            ("v_campanario_inicial", v_campanario[0] if v_campanario else 0),
            ("v_campanario_final", v_campanario[-1] if v_campanario else 0),
            ("v_campanario_minimo", min(v_campanario) if v_campanario else 0),
            ("v_campanario_maximo", max(v_campanario) if v_campanario else 0),
            ("cmg_promedio", sum(cmg) / len(cmg) if cmg else 0),
            (
                "potencia_ch4_promedio",
                sum(potencia_ch4) / len(potencia_ch4) if potencia_ch4 else 0,
            ),
            (
                "potencia_ch6_promedio",
                sum(potencia_ch6) / len(potencia_ch6) if potencia_ch6 else 0,
            ),
            ("ingreso_total_estimado", ingreso_total_estimado),
        ]

        # Título sección metadatos
        ws_resumen.append(["Metadatos", ""])
        for key, value in metadata_rows:
            ws_resumen.append([key, value])

        ws_resumen.append([])
        ws_resumen.append(["Metricas operativas", ""])
        for key, value in metric_rows:
            ws_resumen.append([key, value])

        # Formato hoja resumen
        section_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill("solid", fgColor="1F4E78")
        label_font = Font(bold=True)

        for row_idx in (1, len(metadata_rows) + 3):
            ws_resumen[f"A{row_idx}"].font = section_font
            ws_resumen[f"A{row_idx}"].fill = section_fill
            ws_resumen[f"B{row_idx}"].fill = section_fill

        for row_idx in range(2, len(metadata_rows) + 2):
            ws_resumen[f"A{row_idx}"].font = label_font

        metrics_start = len(metadata_rows) + 4
        for row_idx in range(
            metrics_start + 1, metrics_start + len(metric_rows) + 1
        ):
            ws_resumen[f"A{row_idx}"].font = label_font

        ws_resumen.column_dimensions["A"].width = 28
        ws_resumen.column_dimensions["B"].width = 24

        for excel_row in ws_resumen.iter_rows(
            min_row=1, max_row=ws_resumen.max_row, min_col=1, max_col=2
        ):
            for cell in excel_row:
                cell.alignment = Alignment(vertical="center")

        # Hoja Resultados
        ws = wb.create_sheet(title="Resultados")

        headers = [
            "periodo",
            "q_opt",
            "v_cincel",
            "v_campanario",
            "cmg",
            "potencia_ch4",
            "potencia_ch6",
            "ingreso",
        ]
        ws.append(headers)

        n = len(q_opt)

        for i in range(n):
            ws.append(
                [
                    i + 1,
                    q_opt[i] if i < len(q_opt) else None,
                    v_cincel[i + 1] if i + 1 < len(v_cincel) else None,
                    v_campanario[i + 1] if i + 1 < len(v_campanario) else None,
                    cmg[i] if i < len(cmg) else None,
                    potencia_ch4[i] if i < len(potencia_ch4) else None,
                    potencia_ch6[i] if i < len(potencia_ch6) else None,
                    ingresos[i] if i < len(ingresos) else None,
                ]
            )

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        ws.freeze_panes = "A2"

        column_widths = {
            1: 12,
            2: 14,
            3: 16,
            4: 18,
            5: 12,
            6: 16,
            7: 16,
            8: 16,
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_cells[0].number_format = "0"
            for cell in row_cells[1:]:
                cell.number_format = "0.0000"

        # Hoja Graficos
        ws_graficos = wb.create_sheet(title="Graficos")

        # Grafico 1: Q opt por periodo
        chart_qopt = LineChart()
        chart_qopt.title = "Q opt por período"
        chart_qopt.y_axis.title = "Q opt"
        chart_qopt.x_axis.title = "Periodo"
        chart_qopt.height = 8
        chart_qopt.width = 16

        data_qopt = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats_periodo = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart_qopt.add_data(data_qopt, titles_from_data=True)
        chart_qopt.set_categories(cats_periodo)
        ws_graficos.add_chart(chart_qopt, "A1")

        # Grafico 2: Volumenes de embalses
        chart_vol = LineChart()
        chart_vol.title = "Volúmenes de embalses"
        chart_vol.y_axis.title = "Volumen"
        chart_vol.x_axis.title = "Periodo"
        chart_vol.height = 8
        chart_vol.width = 16

        data_vol = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=ws.max_row)
        chart_vol.add_data(data_vol, titles_from_data=True)
        chart_vol.set_categories(cats_periodo)
        ws_graficos.add_chart(chart_vol, "A20")

        # Grafico 3: Ingreso por periodo
        chart_ingreso = LineChart()
        chart_ingreso.title = "Ingreso por período"
        chart_ingreso.y_axis.title = "Ingreso"
        chart_ingreso.x_axis.title = "Periodo"
        chart_ingreso.height = 8
        chart_ingreso.width = 16

        data_ingreso = Reference(ws, min_col=8, min_row=1, max_row=ws.max_row)
        chart_ingreso.add_data(data_ingreso, titles_from_data=True)
        chart_ingreso.set_categories(cats_periodo)
        ws_graficos.add_chart(chart_ingreso, "A39")

        # Grafico 4: CMG vs potencia
        chart_cmg_pot = LineChart()
        chart_cmg_pot.title = "CMG vs potencia"
        chart_cmg_pot.y_axis.title = "Valor"
        chart_cmg_pot.x_axis.title = "Periodo"
        chart_cmg_pot.height = 8
        chart_cmg_pot.width = 16

        data_cmg_pot = Reference(ws, min_col=5, max_col=7, min_row=1, max_row=ws.max_row)
        chart_cmg_pot.add_data(data_cmg_pot, titles_from_data=True)
        chart_cmg_pot.set_categories(cats_periodo)
        ws_graficos.add_chart(chart_cmg_pot, "A58")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="corrida_{corrida.id}.xlsx"'
            },
        )
    finally:
        db.close()